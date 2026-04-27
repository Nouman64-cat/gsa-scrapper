import pandas as pd
import os
import re
import sys
import io
import logging
from datetime import datetime
from urllib.parse import urlparse, parse_qs
from sqlmodel import Session, select

# Ensure the root project dir is in sys.path so we can import modules
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from database.db import get_engine
from database.models import GSALink, GSAScrapedData, ImportedLink, ImportedPart, LinkScrapedData

logger = logging.getLogger(__name__)


def _compute_variation(card_pn: str, imported_pn: str) -> str:
    """
    Compare a card's scraped part number against the imported (target) part number.

    Returns:
        "Same"      – exact case-insensitive match
        "Different" – match only after stripping all non-alphanumeric characters
        ""          – no match or either value is empty
    """
    if not card_pn or not imported_pn:
        return ""
    if card_pn.strip().upper() == imported_pn.strip().upper():
        return "Same"
    norm_card     = re.sub(r"[^A-Z0-9]", "", card_pn.upper())
    norm_imported = re.sub(r"[^A-Z0-9]", "", imported_pn.upper())
    if norm_card and norm_imported and norm_card == norm_imported:
        return "Different"
    return ""


def get_export_info() -> dict:
    """
    Return a lightweight summary of what data is available to export.
    Used by the frontend to show the user what will be in their download
    before actually streaming the file.

    Returns:
        {
          "has_parts_data":  bool,  # gsa_scraped_data has records
          "has_links_data":  bool,  # links_scraped_data has records
          "parts_records":   int,
          "links_records":   int,
          "active_engine":   "parts" | "links" | "both" | "none"
        }
    """
    try:
        engine = get_engine()
        with Session(engine) as session:
            parts_records = session.query(GSAScrapedData).count()
            links_records = session.query(LinkScrapedData).count()
    except Exception as e:
        logger.error(f"get_export_info DB error: {e}")
        return {"has_parts_data": False, "has_links_data": False,
                "parts_records": 0, "links_records": 0, "active_engine": "none"}

    has_parts = parts_records > 0
    has_links = links_records > 0

    if has_parts and has_links:
        engine_label = "both"
    elif has_parts:
        engine_label = "parts"
    elif has_links:
        engine_label = "links"
    else:
        engine_label = "none"

    return {
        "has_parts_data": has_parts,
        "has_links_data": has_links,
        "parts_records": parts_records,
        "links_records": links_records,
        "active_engine": engine_label,
    }


def export_to_excel():
    """
    Smart export.

    Sheet decision table
    ─────────────────────────────────────────────────────────────────────────────
    GSA parts data     Internal scraped   External scraped   Sheets written
    ─────────────────  ─────────────────  ─────────────────  ────────────────────
    has records        –                  –                  "GSA Parts Data"
    –                  has records        –                  "Internal Links"
    –                  –                  has records        "External Links"
    –                  has records        has records        both link sheets
    has records        has records        has records        all three sheets
    none               none               none               → None (error)
    ─────────────────────────────────────────────────────────────────────────────

    Internal Links sheet columns (one row per product-detail link):
        Internal Link URL | Manufacturer Part Name |
        GSA PRICE | Unit | Contractor | contract#: |          ← slot 1
        GSA PRICE.1 | Unit.1 | Contractor.1 | contract#:.1 | ← slot 2  … up to slot 6

    External Links sheet columns (one row per search/external link):
        Manufacturer Part Number | External Link URL |
        GSA PRICE | Unit | Manufacturer Part Name | Contractor | contract#: |   ← slot 1
        GSA PRICE.1 | … up to slot 6

    Returns (BytesIO buffer, filename) on success, or None on failure.
    """
    from collections import defaultdict

    try:
        engine = get_engine()
    except Exception as e:
        logger.error(f"Failed to connect to the database: {e}")
        return None

    try:
        with Session(engine) as session:
            imported_parts = session.exec(select(ImportedPart).order_by(ImportedPart.id)).all()
            scraped_parts  = session.exec(select(GSAScrapedData)).all()
            gsa_links_all  = session.exec(select(GSALink)).all()

            # Build per-link lookups from imported_links.
            # link_type_map  : link_id → "internal" | "external"  (used for tab routing)
            # link_import_pn : link_id → imported part_number      (used for Part Variation)
            # current_link_ids is the session boundary: orphan rows from prior sessions
            # whose link_id no longer exists in imported_links are silently excluded.
            imported_links_all = session.exec(select(ImportedLink)).all()
            link_type_map: dict[int, str] = {
                il.id: il.link_type for il in imported_links_all
            }
            link_import_pn: dict[int, str] = {
                il.id: (il.part_number or "") for il in imported_links_all
            }
            current_link_ids: set[int] = set(link_type_map.keys())

            links_scraped_raw = session.exec(
                select(LinkScrapedData).order_by(
                    LinkScrapedData.link_id, LinkScrapedData.row_order
                )
            ).all()

        # Part numbers that were attempted by the scraper (is_scraped=True in gsa_links)
        attempted_part_numbers: set[str] = {
            gl.part_number for gl in gsa_links_all if gl.is_scraped
        }

        # Session filter: drop rows whose link_id is no longer in imported_links
        links_scraped = [r for r in links_scraped_raw if r.link_id in current_link_ids]

        # Split by link type
        internal_scraped = [r for r in links_scraped if link_type_map.get(r.link_id) == "internal"]
        external_scraped = [r for r in links_scraped if link_type_map.get(r.link_id) == "external"]

        # Links that were attempted (is_scraped=True) but produced zero rows
        scraped_link_ids: set[int] = {r.link_id for r in links_scraped}
        failed_internal_links = [
            il for il in imported_links_all
            if il.is_scraped
            and link_type_map.get(il.id) == "internal"
            and il.id not in scraped_link_ids
        ]
        failed_external_links = [
            il for il in imported_links_all
            if il.is_scraped
            and link_type_map.get(il.id) == "external"
            and il.id not in scraped_link_ids
        ]

        has_parts_data    = len(scraped_parts) > 0
        has_internal_data = len(internal_scraped) > 0 or len(failed_internal_links) > 0
        has_external_data = len(external_scraped) > 0 or len(failed_external_links) > 0

        logger.info(
            f"Export info: parts={len(scraped_parts)}, "
            f"internal_links={len(internal_scraped)} (+{len(failed_internal_links)} failed), "
            f"external_links={len(external_scraped)} (+{len(failed_external_links)} failed)"
        )

        if not has_parts_data and not has_internal_data and not has_external_data:
            logger.error("No scraped data found. Run an extraction first.")
            return None

        timestamp = datetime.now().strftime("%Y_%m_%d_%H_%M_%S")
        output_buffer = io.BytesIO()

        # Slot suffixes: "" for slot 1, ".1" … ".5" for slots 2–6
        _MAX_SLOTS = 6
        _SUFFIXES  = [""] + [f".{i}" for i in range(1, _MAX_SLOTS)]

        def _v(val):
            """Return empty string for None / NaN values."""
            if val is None:
                return ""
            if isinstance(val, float) and pd.isna(val):
                return ""
            return val

        def _pivot_internal(scraped_rows: list) -> pd.DataFrame:
            """
            Build the Internal Links sheet.

            Columns: Internal Link URL | Manufacturer Part Name |
                     Manufacturer Part Number |
                     GSA PRICE | Unit | Contractor | contract#:  (×6 slots)

            The Manufacturer Part Number is extracted from the product detail
            page's 'About This Item' specification table and stored once per
            link (same value across all slots for a given product).
            """
            groups: dict = defaultdict(list)
            for r in scraped_rows:
                groups[r.link_id].append(r)

            pivoted = []
            for link_id in sorted(groups.keys()):
                rows = sorted(groups[link_id], key=lambda r: r.row_order)
                base = rows[0]

                record: dict = {
                    "Manufacturer Part Name":   _v(base.manufacturer_part_name),
                    "Manufacturer Part Number": _v(base.manufacturer_part_number),
                }
                for i, sfx in enumerate(_SUFFIXES):
                    if i < len(rows):
                        r = rows[i]
                        record[f"GSA PRICE{sfx}"]  = _v(r.price)
                        record[f"Unit{sfx}"]        = _v(r.unit)
                        record[f"Contractor{sfx}"]  = _v(r.contractor_name)
                        record[f"contract#:{sfx}"]  = _v(r.contract_number)
                    else:
                        record[f"GSA PRICE{sfx}"]  = ""
                        record[f"Unit{sfx}"]        = ""
                        record[f"Contractor{sfx}"]  = ""
                        record[f"contract#:{sfx}"]  = ""
                    record[f"___SEP_{i}"] = ""
                pivoted.append(record)

            return pd.DataFrame(pivoted)

        def _pivot_external(scraped_rows: list, import_pn_map: dict) -> pd.DataFrame:
            """
            Build the External Links sheet.

            Column layout:
              Col A : Imported Part Number  ← reference PN from the uploaded file (yellow)

              Per slot (×6), separated by a black ___SEP_ column:
                Part Variation           (yellow)
                Manufacturer Part Number (yellow)  ← card's scraped PN
                Manufacturer Part Name   (yellow)
                Product Name             (blue)
                GSA PRICE                (blue)
                Unit                     (blue)
                Contractor               (blue)
                contract#:               (blue)

            'Part Variation' is "Same" for an exact case-insensitive match against
            the imported PN, or "Different" when the match requires stripping
            special characters.
            """
            groups: dict = defaultdict(list)
            for r in scraped_rows:
                groups[r.link_id].append(r)

            pivoted = []
            for link_id in sorted(groups.keys()):
                rows = sorted(groups[link_id], key=lambda r: r.row_order)
                base = rows[0]

                imported_pn = import_pn_map.get(link_id, "") or _v(base.manufacturer_part_number)

                # Col A: the reference PN from the uploaded import file
                record: dict = {
                    "Imported Part Number": imported_pn,
                }
                for i, sfx in enumerate(_SUFFIXES):
                    if i < len(rows):
                        r = rows[i]
                        card_pn = _v(r.manufacturer_part_number)
                        # --- yellow columns first ---
                        record[f"Part Variation{sfx}"]           = _compute_variation(card_pn, imported_pn)
                        record[f"Manufacturer Part Number{sfx}"] = card_pn
                        record[f"Manufacturer Part Name{sfx}"]   = _v(r.manufacturer_part_name)
                        # --- blue columns after ---
                        record[f"Product Name{sfx}"]             = _v(r.product_name)
                        record[f"GSA PRICE{sfx}"]               = _v(r.price)
                        record[f"Unit{sfx}"]                     = _v(r.unit)
                        record[f"Contractor{sfx}"]               = _v(r.contractor_name)
                        record[f"contract#:{sfx}"]               = _v(r.contract_number)
                    else:
                        record[f"Part Variation{sfx}"]           = ""
                        record[f"Manufacturer Part Number{sfx}"] = ""
                        record[f"Manufacturer Part Name{sfx}"]   = ""
                        record[f"Product Name{sfx}"]             = ""
                        record[f"GSA PRICE{sfx}"]               = ""
                        record[f"Unit{sfx}"]                     = ""
                        record[f"Contractor{sfx}"]               = ""
                        record[f"contract#:{sfx}"]               = ""
                    record[f"___SEP_{i}"] = ""
                pivoted.append(record)

            return pd.DataFrame(pivoted)

        def _style_header(ws):
            """Apply blue background with bold white text to the header row of a worksheet."""
            from openpyxl.styles import PatternFill, Font, Alignment
            blue_fill  = PatternFill(fill_type="solid", fgColor="1F4E79")
            white_bold = Font(bold=True, color="FFFFFF")
            center     = Alignment(horizontal="center", vertical="center")
            for cell in ws[1]:
                if cell.value and str(cell.value).startswith("___SEP_"):
                    continue
                cell.fill      = blue_fill
                cell.font      = white_bold
                cell.alignment = center

        def _style_header_internal(ws):
            """
            Style the Internal Links header row:
              - First two non-separator columns (A=Manufacturer Part Name,
                B=Manufacturer Part Number) → yellow, bold black text
              - Remaining columns → light-blue, bold black text
              - ___SEP_ columns are skipped (handled by _apply_separator_fill)
            """
            from openpyxl.styles import PatternFill, Font, Alignment
            yellow_fill     = PatternFill(fill_type="solid", fgColor="FFFF00")
            light_blue_fill = PatternFill(fill_type="solid", fgColor="9DC3E6")
            bold_black      = Font(bold=True, color="000000")
            center          = Alignment(horizontal="center", vertical="center")

            non_sep_idx = 0
            for cell in ws[1]:
                if cell.value and str(cell.value).startswith("___SEP_"):
                    continue
                cell.font      = bold_black
                cell.alignment = center
                cell.fill      = yellow_fill if non_sep_idx < 2 else light_blue_fill
                non_sep_idx   += 1

        def _style_header_external(ws):
            """
            Style the External Links header row:
              - Manufacturer Part Name, Manufacturer Part Number, Part Variation
                (base column and all per-slot suffixed variants) → yellow, bold black text
              - All other non-separator columns → light-blue, bold black text
              - ___SEP_ columns are skipped (handled by _apply_separator_fill)
            """
            import re as _re
            from openpyxl.styles import PatternFill, Font, Alignment
            yellow_fill     = PatternFill(fill_type="solid", fgColor="FFFF00")
            light_blue_fill = PatternFill(fill_type="solid", fgColor="9DC3E6")
            bold_black      = Font(bold=True, color="000000")
            center          = Alignment(horizontal="center", vertical="center")
            yellow_cols     = {
                "Imported Part Number",
                "Part Variation",
            }

            for cell in ws[1]:
                if cell.value and str(cell.value).startswith("___SEP_"):
                    continue
                cell.font      = bold_black
                cell.alignment = center
                # Strip slot suffix (.1, .2 …) before comparing
                base_name = _re.sub(r'\.\d+$', '', str(cell.value or ""))
                cell.fill = yellow_fill if base_name in yellow_cols else light_blue_fill

        def _apply_separator_fill(ws):
            """Fill every cell in ___SEP_ columns solid black and clear their values."""
            from openpyxl.styles import PatternFill
            black_fill = PatternFill(fill_type="solid", fgColor="000000")
            sep_col_indices = [
                cell.column for cell in ws[1]
                if cell.value and str(cell.value).startswith("___SEP_")
            ]
            max_row = max(ws.max_row, 1)
            for col_idx in sep_col_indices:
                for row_idx in range(1, max_row + 1):
                    c = ws.cell(row=row_idx, column=col_idx)
                    c.fill  = black_fill
                    c.value = None

        def _highlight_parts_failed_rows(ws, failed_row_indices: list):
            """Highlight specific rows (1-based, where row 1 is header) with red."""
            from openpyxl.styles import PatternFill, Font
            red_fill   = PatternFill(fill_type="solid", fgColor="FF0000")
            white_bold = Font(bold=True, color="FFFFFF")
            for row_idx in failed_row_indices:
                for cell in ws[row_idx]:
                    cell.fill = red_fill
                    cell.font = white_bold

        def _highlight_failed_rows(ws, first_failed_row: int):
            """
            Highlight every non-separator cell from first_failed_row to the last
            row with a red background.  Separator columns are intentionally skipped
            here because _apply_separator_fill (which runs after this) will paint
            them black regardless.
            """
            from openpyxl.styles import PatternFill, Font
            red_fill  = PatternFill(fill_type="solid", fgColor="FF0000")
            white_bold = Font(bold=True, color="FFFFFF")
            sep_col_indices = {
                cell.column for cell in ws[1]
                if cell.value and str(cell.value).startswith("___SEP_")
            }
            for row in ws.iter_rows(min_row=first_failed_row, max_row=ws.max_row):
                for cell in row:
                    if cell.column not in sep_col_indices:
                        cell.fill = red_fill
                        cell.font = white_bold

        def _build_failed_internal_rows(failed_links: list) -> list[dict]:
            """
            Build one record per failed internal link.

            Manufacturer Part Name  ← mfrName query param from the product_detail URL
            Manufacturer Part Number ← itemNumber query param
            All slot columns         ← empty strings
            """
            records = []
            for il in failed_links:
                try:
                    params = parse_qs(urlparse(il.link).query)
                    mfr_pn   = params.get("itemNumber", [""])[0]
                    mfr_name = params.get("mfrName",    [""])[0]
                except Exception:
                    mfr_pn, mfr_name = "", ""
                record: dict = {
                    "Manufacturer Part Name":   mfr_name,
                    "Manufacturer Part Number": mfr_pn,
                }
                for i, sfx in enumerate(_SUFFIXES):
                    record[f"GSA PRICE{sfx}"]  = ""
                    record[f"Unit{sfx}"]        = ""
                    record[f"Contractor{sfx}"]  = ""
                    record[f"contract#:{sfx}"]  = ""
                    record[f"___SEP_{i}"]       = ""
                records.append(record)
            return records

        def _build_failed_external_rows(failed_links: list, import_pn_map: dict) -> list[dict]:
            """
            Build one record per failed external link.

            Imported Part Number ← from the import file (already stored in import_pn_map)
            All slot columns     ← empty strings
            """
            records = []
            for il in failed_links:
                imported_pn = import_pn_map.get(il.id, "") or ""
                record: dict = {"Imported Part Number": imported_pn}
                for i, sfx in enumerate(_SUFFIXES):
                    record[f"Part Variation{sfx}"]           = ""
                    record[f"Manufacturer Part Number{sfx}"] = ""
                    record[f"Manufacturer Part Name{sfx}"]   = ""
                    record[f"Product Name{sfx}"]             = ""
                    record[f"GSA PRICE{sfx}"]               = ""
                    record[f"Unit{sfx}"]                     = ""
                    record[f"Contractor{sfx}"]               = ""
                    record[f"contract#:{sfx}"]               = ""
                    record[f"___SEP_{i}"]                    = ""
                records.append(record)
            return records

        with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:

            # ── GSA Parts Data (price extraction pipeline) ────────────────────
            if has_parts_data:
                rows = [
                    {"part_number": r.part_number, "manufacturer": r.manufacturer or ""}
                    for r in imported_parts
                ]
                df = pd.DataFrame(rows) if rows else pd.DataFrame(
                    columns=["part_number", "manufacturer"]
                )
                for col in ['1 GSA Low Price', 'Unit', 'Contractor:Name',
                            '2 GSA Low Price', 'Unit.1', 'Contractor:Name.1']:
                    df[col] = pd.Series([None] * len(df), dtype='object')

                scraped_dict = {str(s.part_number).strip(): s for s in scraped_parts}
                matched = 0
                # worksheet row index for each df row (header=1, data starts at 2)
                failed_ws_rows: list[int] = []
                for df_pos, (idx, row) in enumerate(df.iterrows()):
                    pn = str(row['part_number']).strip()
                    ws_row = df_pos + 2  # +1 for header, +1 for 1-based index
                    if pn in scraped_dict:
                        s = scraped_dict[pn]
                        df.at[idx, '1 GSA Low Price']   = _v(s.gsa_low_price_1)
                        df.at[idx, 'Unit']               = _v(s.unit_1)
                        df.at[idx, 'Contractor:Name']    = _v(s.contractor_1)
                        df.at[idx, '2 GSA Low Price']    = _v(s.gsa_low_price_2)
                        df.at[idx, 'Unit.1']             = _v(s.unit_2)
                        df.at[idx, 'Contractor:Name.1']  = _v(s.contractor_2)
                        # Contractor-only (no price extracted) → treat as failed
                        if s.gsa_low_price_1 is None and s.gsa_low_price_2 is None:
                            failed_ws_rows.append(ws_row)
                        else:
                            matched += 1
                    elif pn in attempted_part_numbers:
                        # Scraping was attempted but produced no results
                        failed_ws_rows.append(ws_row)

                df.to_excel(writer, sheet_name="GSA Parts Data", index=False)
                ws_parts = writer.sheets["GSA Parts Data"]
                _style_header(ws_parts)
                if failed_ws_rows:
                    _highlight_parts_failed_rows(ws_parts, failed_ws_rows)
                logger.info(
                    f"Export 'GSA Parts Data': {matched} row(s) matched, "
                    f"{len(failed_ws_rows)} row(s) failed (red)"
                )

            # ── Internal Links (product-detail link extraction) ───────────────
            if has_internal_data:
                df_int = _pivot_internal(internal_scraped) if internal_scraped else pd.DataFrame()
                n_int_success = len(df_int)

                failed_int_rows = _build_failed_internal_rows(failed_internal_links)
                if failed_int_rows:
                    df_failed_int = pd.DataFrame(failed_int_rows)
                    df_int = pd.concat([df_int, df_failed_int], ignore_index=True) \
                             if n_int_success > 0 else df_failed_int

                df_int.to_excel(writer, sheet_name="Internal Links", index=False)
                ws_int = writer.sheets["Internal Links"]
                _style_header_internal(ws_int)
                if failed_int_rows:
                    _highlight_failed_rows(ws_int, first_failed_row=n_int_success + 2)
                _apply_separator_fill(ws_int)
                logger.info(
                    f"Export 'Internal Links': {n_int_success} success, "
                    f"{len(failed_int_rows)} failed (red rows)"
                )

            # ── External Links (search/external link extraction) ──────────────
            if has_external_data:
                df_ext = _pivot_external(external_scraped, link_import_pn) \
                         if external_scraped else pd.DataFrame()
                n_ext_success = len(df_ext)

                failed_ext_rows = _build_failed_external_rows(failed_external_links, link_import_pn)
                if failed_ext_rows:
                    df_failed_ext = pd.DataFrame(failed_ext_rows)
                    df_ext = pd.concat([df_ext, df_failed_ext], ignore_index=True) \
                             if n_ext_success > 0 else df_failed_ext

                df_ext.to_excel(writer, sheet_name="External Links", index=False)
                ws_ext = writer.sheets["External Links"]
                _style_header_external(ws_ext)
                if failed_ext_rows:
                    _highlight_failed_rows(ws_ext, first_failed_row=n_ext_success + 2)
                _apply_separator_fill(ws_ext)
                logger.info(
                    f"Export 'External Links': {n_ext_success} success, "
                    f"{len(failed_ext_rows)} failed (red rows)"
                )

        # ── Filename ──────────────────────────────────────────────────────────
        parts_tag    = "parts_"    if has_parts_data    else ""
        internal_tag = "internal_" if has_internal_data else ""
        external_tag = "external_" if has_external_data else ""
        filename = f"gsa_{parts_tag}{internal_tag}{external_tag}export_{timestamp}.xlsx"

        output_buffer.seek(0)
        logger.info(f"Export ready: {filename}")
        return output_buffer, filename

    except Exception as e:
        logger.error(f"Export failed: {e}", exc_info=True)
        return None
